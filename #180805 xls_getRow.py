import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string

wb = openpyxl.load_workbook('example.xlsx')

allSht = wb.sheetnames

sheet1 = wb['데이터']

#번호로 열의 문자 알기
_1 = '번호로 열의 문자 알기-50'
print("{}{}".format("1.",_1),end='\t')
a = get_column_letter(50)
print(a)

#열의 문자로 숫자 알아내기
_2 = '열의 문자로 숫자 알아내기-C '
print("{}{}".format("2.",_2),end='\t')
b = column_index_from_string('c')
print(b)

#활용하기
_3 = '번호로 열의 문자 알아내기 활용하기'
print("{}{}".format("3. ",_3),end='\t')
c = get_column_letter(sheet1.max_column)
print(c)

