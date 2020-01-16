import openpyxl

wb = openpyxl.load_workbook('example.xlsx')

#엑셀의 모든시트 이름 알아보기
_1 = '엑셀의 모든시트 이름 출력'
all_sheet = wb.sheetnames
print(_1,'\n',all_sheet,'\n')

#엑셀 시트에 별명붙이기
_2 = '엑셀 시트에 별명붙이기'
sheet1 = wb['데이터']
print(_2)
print('1.', sheet1)
print('2.',sheet1['a1'],'\n')

#엑셀 셀의 값 출력하기
_3 = 'a1의 값 출력하기'
print(_3)
print('3.',sheet1['a1'].value,'\n')

#행번호가 홀수인 값 출력하기
_4='행번호가 홀수인 값 출력하기'
_b1 = sheet1.cell(row = 1, column = 2)
print(_4)
print('1.',_b1)
print('2',_b1.value)
print('3','홀수값 출력')
for i in range(1, 50, 2):
    print("  ",i, sheet1.cell(row = i, column = 3).value)
print('')

#엑셀 행과 열의 최고값 알아내기
_5 = '행과 열의 최고값 알아내기'
lastRow = sheet1.max_row #마지막 행의 값 알아내기
lastCol = sheet1.max_column #마지막 열의 값 알아내기
print(_5)
print('{} {}'.format('마지막 행 :',lastRow)) #양식 및 형식에 맞게 출력하기
print('{} {}'.format('마지막 열 :',lastCol))

# _a = print("{}{} {}".format(idx, '위', s)) # 출력하기 - 양식에 맞게



