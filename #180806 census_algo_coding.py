#개요
'''
1. 프로그램의 목적 : 각 카운티의 통계 산출 하기
2. 프로그램이 하는 일
    - 엑셀 스프레드 시트에서 데이터 읽기
    - 각 카운티의 인구조사 지역 단위의 수를 계산
    - 각 카운티의 총 인구 계산
    - 결과를 출력

3. 코드의 할일
    - openpyxl 모듈로 엑셀 문서를 열고 각 셀 읽기
    - 모든 지역 단위와 인구 데이터를 계산하고 이를 데이터 구조에 저장
    - pprint모듈을 사용하여 데이터 구조를 .py 확장자를 가진 텍스트 파일에 저장
'''

#알고리즘

#준비하기 단계
'''
    #준비하기 단계
        openpyxl, pprint 불러오기
        '워크북을 여는중 안내 메세지' 출력
        census 엑셀파일 불러오기
        popula... 시트를 shteet로 지정하기
        듀플 선언 - countyData
'''
import openpyxl, pprint
print('Opening Workbook....')
wb = openpyxl.load_workbook('census.xlsx')
allSht = wb.sheetnames
sheet = wb['Population by Census Tract']
countyData = {}

# 각 카운티의 인구와 트랙을 countyData에 채우기
'''
    # 각 카운티의 인구와 트랙을 countyData에 채우기
        행 반복하기 - 2 부터 마지막 행+1까지
            state는 각 B열의 값을 저장. sheet['B' + str(row)].vaue
            county는 각 C열의 값을 저장.
            pop는 각 D열의 값을 저장.

        # state 키가 존재하는 지 확인하기
        state가 없으면 {}로 초기화 하기
        # county 키가 존재하는 지 확인하기
        county가 없으면 {}로 초기화 하기

        카운티의 수를 세기 위해서 tract값을 1씩 증가시키기

        카운티의 인구를 합산하기
'''

print('Reading Row......')
for row in range(2,sheet.max_row + 1):
    state = sheet['B' + str(row)].value
    county = sheet['C' + str(row)].value
    pop = sheet['D' + str(row)].value

    #자료구조를 설정
    countyData.setdefault(state,{})
    countyData[state].setdefault(county,{'tract':0,'pop':0})

    countyData[state][county]['tract'] += 1
    countyData[state][county]['pop'] += int(pop)

# 결과를 파일로 만들기
'''
    # 결과를 파일로 만들기
        #텍스트 파일을 열고 countyData의 내용 저장하기
        '결과값 쓰기' 출력
        resultFile에 census.py파일 열기
        countyData의 결과값을 Pretty Print함수로 allData에 쓰기
        resultFile 닫기
        '끝' 출력
'''
print('Writing resultes......')
resultFile = open('census.py','w')
resultFile.write('alldata = ' + pprint.pformat(countyData))
resultFile.close()
print('Done')

#테스트
    # a = sheet.max_row
    # print(a)
