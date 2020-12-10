from openpyxl import Workbook
from random import *

wb = Workbook()
ws = wb.active

# 데이터 입력하기
ws.append(["번호","수학","영어"])

for i in range(1, 11): # 10개의 데이터 넣기
		# 1줄에 번호, 0에서 100까지의 점수 랜점으로 입력
		ws.append([i, randint(0, 100), randint(0, 100)]) 

# 1번재 row만 가지고 오기
row_title = ws[1]

# 1번째 줄인 title을 제외하고 2번재 줄에서 6번째 줄까지 가지고 오기
row_range = ws[2:6]

# 전체 row
for row in tuple(ws.rows):
		print(row[2].value)

# 전체 column
for col in tuple(ws.columns):
		print(col[0].value)

# 전체 row
for row in ws.iter_rows():
		print(row[2].value)

# 전체 columns
for col in ws.iter_columns():
		print(col[0].value)

# 2번째 줄부터 11번째 줄까지, 2번째 열부터 3번재 열까지
for row in ws.iter_rows(min_row = 2, max_row = 11, min_col = 2, max_col = 3):
		print(row[0].value, row[1].value) # 수학, 영어
		print(row)

ws.save("sample.xlsx")
