#!/usr/bin/env python3
# -*- coding: utf-8 -*-
import pandas as pd
import numpy as np
import openpyxl
import xlsxwriter


#creat example xlsx file
df = pd.DataFrame(np.random.randn(10,3), columns=list('ABC'))

print('엑셀 파일 생성중..')
writer = pd.ExcelWriter('test.xlsx', engine='xlsxwriter')


print('엑셀 시트 생성중..')
df.to_excel(writer, sheet_name='test_1', index=False)
df.to_excel(writer, sheet_name='test_2', index=False)
df.to_excel(writer, sheet_name='test_3', index=False)
wb = writer.book
ws = writer.sheets['test_2']

print('엑셀 파일 저장중..')
writer.close()

#엑셀 데이터 복사하기
wb = openpyxl.load_workbook('test.xlsx')
ws_name_wanted = 'test_2'
list_all_ws = wb.sheetnames #엑셀의 모든 시트 리스트로 읽기

#test_2시트가 아닐경우 모두 삭제하는 반복문
for item in list_all_ws:
    if item != ws_name_wanted:
        print('%s 시트 삭제..' % (item))
        remove = wb[item]
        wb.remove(remove)


print('%s 시트 선택..' % (ws_name_wanted))
ws = wb['%s' % (ws_name_wanted)] #원하는 워크시트 선택

print('데이터 복사중..')
for row in ws.iter_rows():
    for cell in row:
        print(cell.value)
#         cell_value = cell.value
#         new_col_loc = (chr(int(ord(cell.coordinate[0:1]))+ 4))  #coordinate[0:1]은 'A1'중에서 'A'를 선택 [0:1]은 처음부터 1개의 리스트를 의미. 영어 A를 숫자로 바꿔서 4를 더한 값을 다시 문자로 바꿈. A열에서 5칸 더한다는 의미
#         new_row_loc = cell.coordinate[1:]                       #coordinate[0:1]은 'A1'중에서 '1'를 선택 [1:]은 처음부터 1부터 마지막까지를 말함. 하지만 리스트에는 'A'와 '1'만 있으므로 1을 의미
#         ws['%s%d' % (new_col_loc, int(new_row_loc) + 3 )] = cell_value
#         ws['%s' % (cell.coordinate)] = ' '                     #좌표값을 비워서 다음 셀의 좌표값이 들어가도록 초기화
#
# print('파일 저장완료')
# wb.save('workOut.xlsx')


