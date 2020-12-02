# 파일 불러오기
from openpyxl import load_workbook

wb = load_workbook("sample.xlsx")
ws = wb.active

# 셀의 내용 화면에 찍어보기
for x in range(1, ws.max_row + 1):  # 마지막 행까지
    for y in range(1, ws.max_column + 1):   # 마지마 칼럼까지
        print(ws.cell(row=x, column = y).value, end = '')

    print()
