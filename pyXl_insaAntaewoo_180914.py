import openpyxl, pprint
print('Opening Workbook....')
wb = openpyxl.load_workbook('안산양지초_나교사.xlsx', read_only=True)
allSht = wb.sheetnames
sheet1 = wb['DATA'] #데이터 시트를 sheet로 지정
personData = {} #인사데이터 자료

#시트의 마지막열, 마지막 칼럼 세기 / todo 실제 데이터가 들어 있는 부분만 확인하는 방법 해결
lastRow = sheet1.max_row
lastCol = sheet1.max_column


#데이터 부분 복사하기

#대상 파일 열기

#시트에 데이터 붙여넣기