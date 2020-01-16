#개요
'''
1. 목적 : 각 물품의 판매수량과 가격의 구하기
2. 프로그램이 하는 일
    - 엑셀 스프레드 시트에서 데이터 읽기
    - 각 물품의 단위수 계산
    - 각 물품의 판매량과 판매가격 계산하기
    - 결과를 출력

3. 코드의 할일
    - openpyxl 모듈로 파일을 열고 각 셀 읽기
    - 모든 아이템의 단위와 판매량, 판매가격을 계산하고 이를 데이터 구조에 저장하기
    - pprint모듈을 이용하여 데이터 구조를 .py 확장자를 가진 텍스트 파일에 저장하기
'''

#알고리즘

#준비하기 단계
'''
- openpyxl 임포트
- 안내메세지('Opening workbook')
- 파일 열기 
- 시트 이름 가져오기
- 튜블 설정 produceData
'''
import openpyxl, pprint
print('Opening workbook......')
wb = openpyxl.load_workbook('produceSales.xlsx',data_only=True)
# wb.sheetnames
sheet = wb['Sheet']
produceData = {}

# 각 아이템의 트랙, 판매량, 가격을 produceData에 채우기
'''
- 안내메세지 출력 ('Reading row......')
- 행 반복하기 : 2열부터 마지막 행 +1까지
    item = sheet['A' + str(row)]
    pound = sheet['C' + str(row)]
    total = sheet['D' + str(row)]
    
    item의 트랙, 판매량, 가격을 0으로 초기화 하기
    
    아이템의 단위를 1씩 증가
    아이템의 판매량을 누적 합산하기
    아이템의 가격을 누적 합산하기 
'''
print('Reading row......')
# a = sheet.max_row
for row in range(2, sheet.max_row + 1):
    item = sheet['A' + str(row)].value
    pound = sheet['C' + str(row)].value
    total = sheet['D' + str(row)].value
    produceData.setdefault(item,{'tract':0,'pound':0,'total':0})

    produceData[item]['tract'] += 1
    produceData[item]['pound'] += float(pound)
    produceData[item]['total'] += float(total)

# print(produceData)

#결과를 파일로 저장하기
'''
- 안내메세지 출력('Writing resultes......')
- produceSales.py 파일 생성하고 resultFile로 연결
- pprint모듈을 이용하여 파일 저장하기
- 파일 닫기
- 안내메세지 출력('Done')
'''

print('Wrting resulte......')
resultFile = open('produceSales.py','w')
resultFile.write('alldata : ' + pprint.pformat(produceData))
resultFile.close()
print('Done')