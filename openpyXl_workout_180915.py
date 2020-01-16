#!/usr/bin/env python3
# -*- coding: utf-8 -*-
import pandas as pd
import numpy as np
import openpyxl
import xlsxwriter


#creat example xlsx file
df = pd.DataFrame(np.random.randn(10,3), columns=list('ABC'))

print('엑셀 파일 생성중..')
writer = pd.ExcelWriter('test.xlsx', engine='xlsxwriter')


print('엑셀 시트 생성중..')
df.to_excel(writer, sheet_name='test_1', index=False)
df.to_excel(writer, sheet_name='test_2', index=False)
df.to_excel(writer, sheet_name='test_3', index=False)
wb = writer.book
ws = writer.sheets['test_2']

print('엑셀 파일 저장중..')
writer.close()

#엑셀 데이터 복사하기
wb = openpyxl.load_workbook('test.xlsx')
ws_name_wanted = 'test_2'
list_all_ws = wb.sheetnames #엑셀의 모든 시트 리스트로 읽기

#test_2시트가 아닐경우 모두 삭제하는 반복문
for item in list_all_ws:
    if item != ws_name_wanted:
        print('%s 시트 삭제..' % (item))
        remove = wb[item]
        wb.remove(remove)


print('%s 시트 선택..' % (ws_name_wanted))
ws = wb['%s' % (ws_name_wanted)] #원하는 워크시트 선택

print('데이터 복사중..')
for row in ws.iter_rows():
    for cell in row:
        cell_value = cell.value
        new_col_loc = (chr(int(ord(cell.coordinate[0:1]))+ 4))  #todo 코드 해석
        new_row_loc = cell.coordinate[1:]                       #todo 코드 해석
        ws['%s%d' % (new_col_loc, int(new_row_loc) + 3 )] = cell_value
        ws['%s' % (cell.coordinate)] = '  '

print('파일 저장완료')
wb.save('workOut.xlsx')


