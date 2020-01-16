import openpyxl

#엑셀 파일 열기
wb = openpyxl.load_workbook("example.xlsx")

#엑셀 모든 시트 이름 알아보기
allSht = wb.sheetnames


# todo 시트에 자동으로 번호 붙이기
# sht = []
# for sht in allSht:
#     sht[i] = wb[sht]
#     print(sht)

# for i, sht in enumerate(allSht,1):
#     sht[i] = wb[sht]
#     sheet + i = sht[i]
#     print(sht[i])

## 참고 enumerate
    # for idx, title in enumerate(title_list,1): #넘버링, 순위매기기
    #     _a = print("{} {}{}{}".format('자료명',idx,'.',title.text.replace('\n','').strip()),file=out) #출력하기 - 폼에 맞게
    # out.close()

#데이터 시트를 sheet1로 지정하기
sheet1 = wb['데이터']

#마지막 행 알아보기
lastRow = sheet1.max_row
lastCol = sheet1.max_column

#b열의 마지막 값을 알아보기 but lastRow의 값과 결과가 같음.
#todo a열의 실제 마지막 값이 있는 행을 알고 싶으면?
lastCell = len(sheet1['b'])

print(lastCell)

#시트1의 내용을 값을 출력하기
    # a1 = sheet1.cell(1,1).value
    # b1 = sheet1.cell(1,2).value
    # c1 = sheet1.cell(1,3).value
    # print(a1,b1,c1)

for j in range(1, lastRow):
    for i in range(1, lastCol+1):
        s = sheet1.cell(j, i).value
        print(s,end='\t'*2) #옆으로 출력하기, 가로로 출력하기
    print("")
