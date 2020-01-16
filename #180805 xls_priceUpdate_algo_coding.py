#개요
'''
1. 프로그램의 목적 : 아이템 3개의 가격을 수정하기
2. 프로그램 수행 작업
    - 루프로 모든 행을 차례로 거쳐가기
    - 행이 마늘, 샐러리 또는 레몬이라면 가격을 고치기
3. 코드작성
    - 스프레드 시트 파일 열기
    - 각행에 대해 열 A값이 샐러리, 마늘, 레몬인지 검사
    - 위의 경우에 해당되면 열 B의 가격을 고치기
    - 스프레드시트를 새로운 파일에 저장하기
'''

#알고리즘

#준비하기
'''
openxlpy 임포트

파일열기
시트이름 알아보기
시트지정하기

'''
import openpyxl

print('Opening workbook......')
wb = openpyxl.load_workbook('produceSales.xlsx')
# a = wb.sheetnames
sheet = wb['Sheet']

#엄데이트할 물건과 가격을 튜플로 만들기
price_update = {
    'Garlic':3.07,
    'Celery':1.19,
    'Lemon':1.27
}
#행을 반복하고 가격을 업데이트 하기
'''
반복문 : 2부터 마지막 행까지
    행의 1열 물품이름 알아내기
    조건문 : 물품이 업데이트 대상에 있으면
        업데이트 대상의 가격을 행의 2열에 넣기 
'''
print('Reading row......')
i = 0
for rowNum in range(2, sheet.max_row + 1):
    produceName = sheet.cell(row = rowNum, column = 1).value #셀의 값을 변수로 지정하기
    if produceName in price_update: #튜플의 값에 변수의 값이 있으면 for문의 형식과 비슷함.
        sheet.cell(row=rowNum, column=2).value = price_update[produceName] #튜플의 물품의 키값에 해당하는 가격을 셀에 저장
        i += 1
print('Changed for ' + str(i)) #개수 카운트하기
#updateProductSales로 새이름으로 저장하기
wb.save('updateProductSales.xlsx') #새이름 저장하기
print('Done')
